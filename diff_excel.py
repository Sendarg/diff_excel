#!/usr/bin/env python
# -*- coding: utf-8 -*-
# version 1.0 update by le @ 2017.6.10 for qxq
# version 2.0 update by le @ 2017.6.18 for qxq, support group compare
# version 2.1 update by le @ 2017.6.21 for qxq, fix mark label

import os, glob
from openpyxl import Workbook, load_workbook
import argparse
from argparse import RawTextHelpFormatter

old_label = u""
new_label = u""
both_label = u""


def get_unique_key_seqs(Sheet, Keys):
	# get all column value as row's unique key
	K1 = []
	for l in Sheet.rows:
		keys = [Keys]
		if "+" in Keys:
			keys = Keys.split("+")
		k1 = ""
		for k in keys:
			i = ord(k) - 65  # change to number index
			if k1:
				k1 += "^^"
			k1 += unicode(l[i].value)
		K1.append(k1)
	print "== Use Key:\t[ %s ]\tof Sheet:\t[%s]" % (K1[0], Sheet.title)
	return K1


def get_num_key(Keys):
	# turn A+B+C to [0,1,2]
	keys = [Keys]
	if "+" in Keys:
		keys = Keys.split("+")
	Key = [ord(k) - 65 for k in keys]  # change to number index
	return Key


def mark_index_label(Workbook, Sheet, Indexs, Basefilename):
	c = Sheet.max_column
	
	Sheet.cell(row=1, column=c + 1).value = u"Changes"
	for i in Indexs:
		Sheet.cell(row=i[0] + 1, column=c + 1).value = i[1]
	
	name = Basefilename.decode('utf-8').split("/")[-1]
	print "== OK:\t[%s]\t[%s]" % (name, Sheet.title)
	Workbook.save("mark-%s" % name)


def single_workbook(Sheet, Indexs, Key, MarkLabel, Basefilename):
	# store data 2 single workbook not all other data
	wb = Workbook()
	ws = wb.active
	ws.title = Sheet.title
	for i in Indexs:
		for line in Sheet.iter_rows(min_row=i + 1, max_row=i + 1):
			ws.append([c.value for c in line])
	name = Basefilename.decode('utf-8').split("/")[-1]
	print "== OK:\t[%s]\t[%s]" % (name, Sheet.title)
	wb.save("single-%s-%s" % (MarkLabel, name))


def single_unique_workbook(SheetName, OnlyData, MarkLabel, Basefilename):
	wb = Workbook()
	ws = wb.active
	ws.title = SheetName
	
	for d in set(OnlyData):  # unique
		line = d.split("^^")
		ws.append(line)
	name = Basefilename.decode('utf-8').split("/")[-1]
	print "== OK:\t[%s]\t[%s]" % (name, SheetName)
	wb.save("single-unique-%s-%s" % (MarkLabel, name))


def get_sub_index_index(Big_2LevelList, Sub_1LevelList):
	# get all index form K1s seq not just first one
	index_a = [[[i, j] for i, x in enumerate(Big_2LevelList) for j, y in enumerate(Big_2LevelList[i]) if y == o] for o
	           in Sub_1LevelList]
	index_b = [j for i in index_a for j in i]
	return index_b


def group_list_by_index(index, GroupIndex):
	# group 2 single list
	## method1,lambda ?? what mean?
	index_c = [[i for i in index if i[GroupIndex] == u] for u in set(map(lambda x: x[GroupIndex], index))]
	# ## method2 , sort+group
	# index1b.sort(key=itemgetter(0))
	# index1c=[[i for i in g] for k,g in groupby(index1b,key=itemgetter(0))]
	return index_c


def get_sub_index_value(Big_2LevelList, Sub_1LevelList):
	# get all index form K1s seq not just first one
	index_a = [[[i, o] for i, x in enumerate(Big_2LevelList) for j, y in enumerate(Big_2LevelList[i]) if y == o] for o
	           in Sub_1LevelList]
	index_b = [j for i in index_a for j in i]
	return index_b


def diffPlus(Old_Files, New_Files, Sheet_Index_From_Old, Sheet_Index_From_New, Key_From_Old, Key_From_New, Single_File,
             Only_Unique_Key):
	W1s = [load_workbook(filename=old) for old in Old_Files]
	S1s = [w1.worksheets[Sheet_Index_From_Old] for w1 in W1s]
	K1s = [get_unique_key_seqs(s1, Key_From_Old) for s1 in S1s]
	K1 = [j for i in K1s for j in i[1:]]  # 去掉每一个标题
	
	W2s = [load_workbook(filename=new) for new in New_Files]
	S2s = [w2.worksheets[Sheet_Index_From_New] for w2 in W2s]
	K2s = [get_unique_key_seqs(s2, Key_From_New) for s2 in S2s]
	K2 = [j for i in K2s for j in i[1:]]
	
	only1 = set(K1) - set(K2)
	only2 = set(K2) - set(K1)
	both = set(K1) & set(K2)
	
	## only old
	index1 = get_sub_index_index(K1s, only1)
	[[i.append(old_label)] for i in index1]
	both1 = get_sub_index_index(K1s, both)
	[[b.append(both_label)] for b in both1]
	index1_value = get_sub_index_value(K1s, only1)
	
	## only new
	index2 = get_sub_index_index(K2s, only2)
	for i in index2:i.append(new_label)
	both2 = get_sub_index_index(K2s, both)
	[[b.append(both_label)] for b in both2]
	index2_value = get_sub_index_value(K2s, only2)
	
	if not Single_File:
		# clean
		[os.remove(f) for f in glob.glob("mark-*")]
		index1_both1 = group_list_by_index(index1 + both1, 0)
		[mark_index_label(W1s[i[0][0]], S1s[i[0][0]], [j[1:3] for j in i], Old_Files[i[0][0]].name)
		 for i in index1_both1]
		index2_both2 = group_list_by_index(index2 + both2, 0)
		[mark_index_label(W2s[i[0][0]], S2s[i[0][0]], [j[1:3] for j in i], New_Files[i[0][0]].name)
		 for i in index2_both2]
	
	else:
		if not Only_Unique_Key:
			[os.remove(f) for f in glob.glob("single-*")]
			# disable only old, not needed and much write operate
			# [single_workbook(S1s[i[0][0]], [j[1] for j in i], Key_From_Old, old_label, Old_Files[i[0][0]].name) for i in index1c]
			[single_workbook(S2s[i[0][0]], [j[1] for j in i], Key_From_New, new_label, New_Files[i[0][0]].name) for i in
			 group_list_by_index(index2, 0)]
		# single_workbook(s2, both_index2, Key_From_New, u"both", New_File)
		else:
			[os.remove(f) for f in glob.glob("single-unique-*")]
			[single_unique_workbook(S1s[i[0][0]].title, [j[1] for j in i], old_label, Old_Files[i[0][0]].name) for i in
			 group_list_by_index(index1_value, 0)]
			[single_unique_workbook(S2s[i[0][0]].title, [j[1] for j in i], new_label, New_Files[i[0][0]].name) for i in
			 group_list_by_index(index2_value, 0)]
	print "== Done"


def diff1(Old_File, New_File, Sheet_Index_From_Old, Sheet_Index_From_New, Key_From_Old, Key_From_New, Single_File,
          Only_Unique_Key):
	w1 = load_workbook(filename=Old_File)  # data_only=True, copied excel need format&style or not ?
	s1 = w1.worksheets[Sheet_Index_From_Old]
	K1 = get_unique_key_seqs(s1, Key_From_Old)
	
	w2 = load_workbook(filename=New_File)
	s2 = w2.worksheets[Sheet_Index_From_New]
	K2 = get_unique_key_seqs(s2, Key_From_New)
	
	only1 = set(K1) - set(K2)
	only2 = set(K2) - set(K1)
	# both = (set(K1) & set(K2)) - set(K1[0:1]) - set(K2[0:1])  # 去掉标题
	
	# get all index form K1 seq not just first one
	index1 = [[i for i, x in enumerate(K1) if x == o] for o in only1]
	index1 = [j for i in index1 for j in i]
	
	index2 = [[i for i, x in enumerate(K2) if x == o] for o in only2]
	index2 = [j for i in index2 for j in i]
	
	# ## temp remove both for perf
	# both_index1 = [[i for i, x in enumerate(K1) if x == b] for b in both]
	# both_index1 = [j for i in both_index1 for j in i]
	# # same data have different index in 2 files
	# both_index2 = [[i for i, x in enumerate(K2) if x == b] for b in both]
	# both_index2 = [j for i in both_index2 for j in i]
	
	if not Single_File:
		mark_index_label(w1, s1, index1, old_label, Old_File.name)
		mark_index_label(w2, s2, index2, new_label, New_File.name)
	else:
		if not Only_Unique_Key:
			single_workbook(s1, index1, Key_From_Old, old_label, Old_File.name)
			# single_workbook(s1, both_index1, Key_From_Old, "both", Old_File)
			single_workbook(s2, index2, Key_From_New, new_label, New_File.name)
		# single_workbook(s2, both_index2, Key_From_New, u"both", New_File)
		else:
			single_unique_workbook(s1.title, only1, old_label, Old_File.name)
			single_unique_workbook(s2.title, only2, new_label, New_File.name)
	print "== Done"


if __name__ == '__main__':
	parser = argparse.ArgumentParser(description='''
    Compare 2 Excel groups in line level by set columns as unique KEY.
    Output0: Mark a user defined label to a new last column in new copied excel files named mark-*.
    Output1: Different lines to single files from them.
    Output2: Different and Unique KEY to single files from them.
    ''', formatter_class=RawTextHelpFormatter)
	
	parser.add_argument("-1", dest="Old_Files", nargs="+", required=True, type=file,
	                    help="older excels to compare, many excels must special same sheet and Keys")
	parser.add_argument("-2", dest="New_Files", nargs="+", required=True, type=file,
	                    help="newer excels to compare, many excels must special same sheet and Keys")
	parser.add_argument("-s1", dest="Sheet_Index_From_Old", type=int, default=0, choices=[0, 1, 2, 3, 4, 5, 6],
	                    help="sheet index need to compare from older excel, default is 0")
	parser.add_argument("-s2", dest="Sheet_Index_From_New", type=int, default=0, choices=[0, 1, 2, 3, 4, 5, 6],
	                    help="sheet index need to compare from newer excel, default is 0")
	parser.add_argument("-k1", dest="Key_From_Old", required=True,
	                    help="unique Key Column from older excel,Simple: A or A+C or B+D+G")
	parser.add_argument("-k2", dest="Key_From_New", required=True,
	                    help="unique Key Column from newer excel,Simple: A or B+C or D+A+F")
	parser.add_argument("-l0", dest="Mark_Both_Label", default=u"both",
	                    help="Special a word to mark both line in all files as Label, default is \"both\"")
	parser.add_argument("-l1", dest="Mark_Old_Label", default=u"old",
	                    help="Special a word to mark different line in older files as Label, default is \"old\"")
	parser.add_argument("-l2", dest="Mark_New_Label", default=u"new",
	                    help="Special a word to mark different line in newer files as Label, default is \"new\"")
	parser.add_argument("-s", dest="Single_File", action="store_true", default=False,
	                    help="output different lines to 2 single excel file")
	parser.add_argument("-u", dest="Only_Unique_Key", action="store_true", default=False,
	                    help="only output unique different key column data to 2 single excel file, use only with -s option")
	parser.add_argument('-v', action='version', version='%(prog)s 2.1')
	args = parser.parse_args()
	
	old_label = args.Mark_Old_Label
	new_label = args.Mark_New_Label
	both_label = args.Mark_Both_Label
	
	diffPlus(args.Old_Files, args.New_Files, args.Sheet_Index_From_Old, args.Sheet_Index_From_New, args.Key_From_Old,
	         args.Key_From_New, args.Single_File, args.Only_Unique_Key)
